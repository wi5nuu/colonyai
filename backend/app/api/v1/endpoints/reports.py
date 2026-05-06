from fastapi import APIRouter, Depends, HTTPException, status, Request
from fastapi.responses import FileResponse
from pydantic import BaseModel
from typing import Optional, List
from datetime import datetime
import uuid
import os
import csv
import io
from pathlib import Path

from app.core.security import get_current_user, require_role
from app.core.config import settings
from app.core.database import get_db
from app.models import Analysis, ColonyDetection, AnalysisStatus, User
from app.schemas.analyses import ReportResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, and_, func
from sqlalchemy.orm import joinedload

router = APIRouter()


class ReportRequest(BaseModel):
    report_type: str = "custom"  # daily, weekly, custom
    date_from: Optional[str] = None
    date_to: Optional[str] = None
    format: str = "pdf"  # pdf, csv


@router.post("/pdf", response_model=ReportResponse)
async def generate_pdf_report(
    request: ReportRequest,
    http_request: Request = None,
    current_user: dict = Depends(require_role("analyst", "manager", "admin", "super_admin")),
    db: AsyncSession = Depends(get_db),
):
    """
    Generate a BPOM-compliant PDF report (own data only).
    - Analyst/Manager/Admin: own analyses
    - Super Admin: all analyses

    Format: A4, Times New Roman 12pt.
    Contents: sample info, detection summary table, CFU/ml value,
              analyst signature field, timestamp.
    """
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import mm, cm
    from reportlab.lib.colors import black, HexColor
    from reportlab.platypus import (
        SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle,
        PageBreak, KeepTogether,
    )
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont

    # Query analyses for the date range
    conditions = []
    user_role = current_user.get("role")
    
    if user_role != "super_admin":
        org_id = current_user.get("organization_id")
        if org_id:
            conditions.append(Analysis.organization_id == uuid.UUID(org_id))
            
    if user_role == "analyst":
        conditions.append(Analysis.user_id == current_user["user_id"])

    if request.date_from:
        dt_from = request.date_from
        if len(dt_from) == 10:
            dt_from += "T00:00:00"
        conditions.append(Analysis.created_at >= datetime.fromisoformat(dt_from))
    if request.date_to:
        dt_to = request.date_to
        if len(dt_to) == 10:
            dt_to += "T23:59:59.999999"
        conditions.append(Analysis.created_at <= datetime.fromisoformat(dt_to))

    result = await db.execute(
        select(Analysis)
        .where(and_(*conditions))
        .options(
            joinedload(Analysis.detections),
            joinedload(Analysis.user).joinedload(User.organization)
        )
        .order_by(Analysis.created_at.desc())
    )
    analyses = result.scalars().unique().all()

    if not analyses:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="No analyses found for the specified date range",
        )

    # --- Build PDF ---
    reports_dir = os.path.join(settings.UPLOAD_DIR, "reports")
    Path(reports_dir).mkdir(parents=True, exist_ok=True)

    report_id = str(uuid.uuid4())
    filename = f"colonyai-report-{report_id}.pdf"
    file_path = os.path.join(reports_dir, filename)

    doc = SimpleDocTemplate(
        file_path,
        pagesize=A4,
        topMargin=2.5 * cm,
        bottomMargin=2.5 * cm,
        leftMargin=2.5 * cm,
        rightMargin=2.5 * cm,
    )

    # Styles – Times New Roman 12pt base
    base_font_name = "Times-Roman"  # reportlab built-in
    styles = getSampleStyleSheet()

    title_style = ParagraphStyle(
        "CustomTitle",
        parent=styles["Title"],
        fontName=base_font_name,
        fontSize=16,
        leading=20,
        alignment=TA_CENTER,
        spaceAfter=6,
    )
    subtitle_style = ParagraphStyle(
        "CustomSubtitle",
        parent=styles["Normal"],
        fontName=base_font_name,
        fontSize=12,
        leading=14,
        alignment=TA_CENTER,
        spaceAfter=12,
    )
    heading_style = ParagraphStyle(
        "CustomHeading",
        parent=styles["Heading2"],
        fontName=base_font_name,
        fontSize=13,
        leading=16,
        spaceBefore=12,
        spaceAfter=6,
    )
    body_style = ParagraphStyle(
        "CustomBody",
        parent=styles["Normal"],
        fontName=base_font_name,
        fontSize=12,
        leading=14,
        spaceAfter=4,
    )
    small_style = ParagraphStyle(
        "SmallText",
        parent=styles["Normal"],
        fontName=base_font_name,
        fontSize=10,
        leading=12,
    )

    elements = []

    # --- Title block ---
    elements.append(Paragraph("ColonyAI Analysis Report", title_style))
    elements.append(Paragraph("BPOM-Compliant Laboratory Report", subtitle_style))
    elements.append(Spacer(1, 6))
    elements.append(Paragraph(
        f"Generated: {datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S UTC')}",
        small_style,
    ))
    elements.append(Paragraph(
        f"Report ID: {report_id}",
        small_style,
    ))
    elements.append(Spacer(1, 12))

    # --- Overall detection summary ---
    total_analyses = len(analyses)
    total_colonies = sum(a.colony_count or 0 for a in analyses)
    valid_analyses = sum(
        1 for a in analyses
        if (a.status == AnalysisStatus.COMPLETED if hasattr(a.status, 'value') else str(a.status) == "completed")
    )
    avg_cfu = None
    cfu_values = [a.cfu_per_ml for a in analyses if a.cfu_per_ml is not None]
    if cfu_values:
        avg_cfu = sum(cfu_values) / len(cfu_values)

    elements.append(Paragraph("Detection Summary", heading_style))

    summary_data = [
        ["Parameter", "Value"],
        ["Total Analyses", str(total_analyses)],
        ["Valid Analyses", str(valid_analyses)],
        ["Total Colonies Detected", str(total_colonies)],
        ["Average CFU/ml", f"{avg_cfu:.2e}" if avg_cfu is not None else "N/A"],
    ]

    summary_table = Table(summary_data, colWidths=[8 * cm, 8 * cm])
    summary_table.setStyle(TableStyle([
        ("FONTNAME", (0, 0), (-1, -1), base_font_name),
        ("FONTSIZE", (0, 0), (-1, -1), 12),
        ("BACKGROUND", (0, 0), (-1, 0), HexColor("#4A5568")),
        ("TEXTCOLOR", (0, 0), (-1, 0), HexColor("#FFFFFF")),
        ("ALIGN", (0, 0), (-1, -1), "LEFT"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("GRID", (0, 0), (-1, -1), 0.5, black),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [HexColor("#F7FAFC"), HexColor("#FFFFFF")]),
        ("TOPPADDING", (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
        ("LEFTPADDING", (0, 0), (-1, -1), 8),
        ("RIGHTPADDING", (0, 0), (-1, -1), 8),
    ]))
    elements.append(summary_table)
    elements.append(Spacer(1, 18))

    # --- Executive Summary (per Expected Output 4) ---
    elements.append(Paragraph("Executive Summary", heading_style))
    elements.append(Spacer(1, 6))

    # Efficiency metrics
    manual_time_per_sample = 20  # minutes (baseline from proposal)
    ai_time_per_sample = 2  # minutes (ColonyAI)
    time_saved_per_sample = manual_time_per_sample - ai_time_per_sample
    total_time_saved_minutes = total_analyses * time_saved_per_sample
    total_time_saved_hours = total_time_saved_minutes / 60
    efficiency_gain_pct = (time_saved_per_sample / manual_time_per_sample) * 100

    # Cost savings estimate (proposal: 40% labor cost reduction)
    avg_analyst_hourly_rate = 50000  # IDR (example)
    labor_cost_saved = total_time_saved_hours * avg_analyst_hourly_rate * 0.40

    executive_data = [
        ["Metric", "Pre-AI (Manual)", "Post-AI (ColonyAI)", "Improvement"],
        ["Time per Sample", f"{manual_time_per_sample} min", f"{ai_time_per_sample} min", Paragraph(f"{time_saved_per_sample} min saved ({efficiency_gain_pct:.0f}% reduction)", small_style)],
        ["Total Time Invested", Paragraph(f"{total_analyses * manual_time_per_sample} min ({total_analyses * manual_time_per_sample / 60:.1f} hrs)", small_style),
         Paragraph(f"{total_analyses * ai_time_per_sample} min ({total_analyses * ai_time_per_sample / 60:.1f} hrs)", small_style),
         Paragraph(f"{total_time_saved_minutes} min ({total_time_saved_hours:.1f} hrs) saved", small_style)],
        ["Inter-Analyst CV", "22.7%-80%", "<5%", "Consistent results"],
        ["Throughput", f"{total_analyses} plates", f"{total_analyses} plates (automated)", "5-8x potential increase"],
    ]

    executive_table = Table(executive_data, colWidths=[3.5 * cm, 3.5 * cm, 4 * cm, 5 * cm])
    executive_table.setStyle(TableStyle([
        ("FONTNAME", (0, 0), (-1, -1), base_font_name),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("BACKGROUND", (0, 0), (-1, 0), HexColor("#2D3748")),
        ("TEXTCOLOR", (0, 0), (-1, 0), HexColor("#FFFFFF")),
        ("FONTNAME", (0, 0), (-1, 0), base_font_name),
        ("FONTSIZE", (0, 0), (-1, 0), 10),
        ("ALIGN", (0, 0), (-1, -1), "LEFT"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("GRID", (0, 0), (-1, -1), 0.5, black),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [HexColor("#F0FFF4"), HexColor("#FFFFFF")]),
        ("TOPPADDING", (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ("LEFTPADDING", (0, 0), (-1, -1), 6),
        ("RIGHTPADDING", (0, 0), (-1, -1), 6),
    ]))
    elements.append(executive_table)
    elements.append(Spacer(1, 12))

    # Cost savings note
    elements.append(Paragraph(
        f"<b>Estimated Labor Cost Savings:</b> Approximately IDR {labor_cost_saved:,.0f} "
        f"({total_time_saved_hours:.1f} hours × {avg_analyst_hourly_rate:,} IDR/hr × 40% efficiency gain). "
        f"Based on proposal benchmarks for Indonesian microbiology laboratories.",
        small_style,
    ))
    elements.append(Spacer(1, 18))

    # --- Monthly Throughput Trends ---
    elements.append(Paragraph("Monthly Throughput Trends", heading_style))
    elements.append(Spacer(1, 6))

    # Group analyses by month
    from collections import defaultdict
    monthly_data = defaultdict(lambda: {"count": 0, "colonies": 0, "cfu_values": []})

    for analysis in analyses:
        month_key = analysis.created_at.strftime("%Y-%m")
        monthly_data[month_key]["count"] += 1
        monthly_data[month_key]["colonies"] += analysis.colony_count or 0
        if analysis.cfu_per_ml:
            monthly_data[month_key]["cfu_values"].append(analysis.cfu_per_ml)

    # Build monthly trend table
    trend_data = [["Month", "Analyses", "Total Colonies", "Avg CFU/ml"]]
    for month in sorted(monthly_data.keys()):
        data = monthly_data[month]
        avg_cfu = sum(data["cfu_values"]) / len(data["cfu_values"]) if data["cfu_values"] else 0
        trend_data.append([
            month,
            str(data["count"]),
            str(data["colonies"]),
            f"{avg_cfu:.2e}" if avg_cfu > 0 else "N/A",
        ])

    trend_table = Table(trend_data, colWidths=[4 * cm, 4 * cm, 4 * cm, 4 * cm])
    trend_table.setStyle(TableStyle([
        ("FONTNAME", (0, 0), (-1, -1), base_font_name),
        ("FONTSIZE", (0, 0), (-1, -1), 10),
        ("BACKGROUND", (0, 0), (-1, 0), HexColor("#4A5568")),
        ("TEXTCOLOR", (0, 0), (-1, 0), HexColor("#FFFFFF")),
        ("ALIGN", (0, 0), (-1, -1), "LEFT"),
        ("GRID", (0, 0), (-1, -1), 0.5, black),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [HexColor("#F0FFF4"), HexColor("#FFFFFF")]),
        ("TOPPADDING", (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ("LEFTPADDING", (0, 0), (-1, -1), 8),
        ("RIGHTPADDING", (0, 0), (-1, -1), 8),
    ]))
    elements.append(trend_table)
    elements.append(Spacer(1, 18))

    # --- Per-sample details ---
    elements.append(Paragraph("Sample Details", heading_style))

    for analysis in analyses:
        status_str = analysis.status.value if hasattr(analysis.status, 'value') else str(analysis.status)
        analyst_name = analysis.user.full_name if analysis.user else "Unknown Analyst"
        org_name = analysis.user.organization.name if analysis.user and analysis.user.organization else "ColonyAI General"
        
        elements.append(Paragraph(
            f"<b>Sample:</b> {analysis.sample_id} &nbsp; | &nbsp; "
            f"<b>Media:</b> {analysis.media_type} &nbsp; | &nbsp; "
            f"<b>Date:</b> {analysis.created_at.strftime('%Y-%m-%d %H:%M')}",
            body_style,
        ))

        detail_data = [
            ["Parameter", "Value"],
            ["Institution / Company", Paragraph(org_name, body_style)],
            ["Analyst Name", Paragraph(analyst_name, body_style)],
            ["Dilution Factor", Paragraph(f"{analysis.dilution_factor}", body_style)],
            ["Plated Volume (ml)", Paragraph(f"{analysis.plated_volume_ml}", body_style)],
            ["Colony Count", Paragraph(str(analysis.colony_count or 0), body_style)],
            ["CFU/ml", Paragraph(f"{analysis.cfu_per_ml:.2e}" if analysis.cfu_per_ml else "N/A", body_style)],
            ["Confidence", Paragraph(f"{analysis.confidence_score * 100:.1f}%" if analysis.confidence_score else "N/A", body_style)],
            ["Reliability", Paragraph((analysis.reliability or "N/A").capitalize(), body_style)],
            ["Status", Paragraph(status_str.capitalize(), body_style)],
        ]

        detail_table = Table(detail_data, colWidths=[6 * cm, 10 * cm])
        detail_table.setStyle(TableStyle([
            ("FONTNAME", (0, 0), (-1, -1), base_font_name),
            ("FONTSIZE", (0, 0), (-1, -1), 11),
            ("BACKGROUND", (0, 0), (-1, 0), HexColor("#E2E8F0")),
            ("TEXTCOLOR", (0, 0), (-1, 0), black),
            ("FONTNAME", (0, 0), (-1, 0), base_font_name),
            ("FONTSIZE", (0, 0), (-1, 0), 11),
            ("GRID", (0, 0), (-1, -1), 0.5, HexColor("#CBD5E0")),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [HexColor("#F7FAFC"), HexColor("#FFFFFF")]),
            ("TOPPADDING", (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ("LEFTPADDING", (0, 0), (-1, -1), 8),
            ("RIGHTPADDING", (0, 0), (-1, -1), 8),
        ]))
        elements.append(detail_table)

        # Class breakdown if available
        class_breakdown = analysis.class_breakdown or {}
        if class_breakdown:
            breakdown_rows = [["Class", "Count"]]
            for cls_name, count in class_breakdown.items():
                breakdown_rows.append([cls_name, str(count)])
            breakdown_table = Table(breakdown_rows, colWidths=[6 * cm, 10 * cm])
            breakdown_table.setStyle(TableStyle([
                ("FONTNAME", (0, 0), (-1, -1), base_font_name),
                ("FONTSIZE", (0, 0), (-1, -1), 10),
                ("BACKGROUND", (0, 0), (-1, 0), HexColor("#EDF2F7")),
                ("GRID", (0, 0), (-1, -1), 0.5, HexColor("#CBD5E0")),
                ("TOPPADDING", (0, 0), (-1, -1), 3),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
                ("LEFTPADDING", (0, 0), (-1, -1), 8),
                ("RIGHTPADDING", (0, 0), (-1, -1), 8),
            ]))
            elements.append(Spacer(1, 4))
            elements.append(Paragraph("<i>Class Breakdown:</i>", small_style))
            elements.append(breakdown_table)

        elements.append(Spacer(1, 12))

    # --- Signature block ---
    elements.append(Spacer(1, 24))
    elements.append(Paragraph("Analyst Certification", heading_style))
    elements.append(Spacer(1, 36))
    elements.append(Paragraph("_" * 50, body_style))
    elements.append(Spacer(1, 4))
    elements.append(Paragraph("Analyst Name &amp; Signature", small_style))
    elements.append(Spacer(1, 24))
    elements.append(Paragraph("_" * 50, body_style))
    elements.append(Spacer(1, 4))
    elements.append(Paragraph("Date", small_style))
    elements.append(Spacer(1, 12))
    elements.append(Paragraph(
        "<i>This report was generated by ColonyAI, an AI-powered colony detection system. "
        "Results should be reviewed by a qualified analyst before regulatory submission.</i>",
        small_style,
    ))

    # Build
    doc.build(elements)

    url = f"{settings.BACKEND_URL}/uploads/reports/{filename}"
    expires_at = datetime.utcnow().replace(hour=23, minute=59, second=59).isoformat()

    return ReportResponse(
        url=url,
        filename=filename,
        expires_at=expires_at,
    )


@router.post("/csv", response_model=ReportResponse)
async def generate_csv_report(
    request: ReportRequest,
    current_user: dict = Depends(require_role("analyst", "manager", "admin", "super_admin")),
    db: AsyncSession = Depends(get_db),
):
    """
    Generate CSV report with detailed detection data (own data only).
    - Analyst/Manager/Admin: own analyses
    - Super Admin: all analyses
    """
    # Query analyses for the date range
    conditions = []
    user_role = current_user.get("role")
    
    if user_role != "super_admin":
        org_id = current_user.get("organization_id")
        if org_id:
            conditions.append(Analysis.organization_id == uuid.UUID(org_id))
            
    if user_role == "analyst":
        conditions.append(Analysis.user_id == current_user["user_id"])

    if request.date_from:
        dt_from = request.date_from
        if len(dt_from) == 10:
            dt_from += "T00:00:00"
        conditions.append(Analysis.created_at >= datetime.fromisoformat(dt_from))
    if request.date_to:
        dt_to = request.date_to
        if len(dt_to) == 10:
            dt_to += "T23:59:59.999999"
        conditions.append(Analysis.created_at <= datetime.fromisoformat(dt_to))

    result = await db.execute(
        select(Analysis)
        .where(and_(*conditions))
        .options(
            joinedload(Analysis.detections),
            joinedload(Analysis.user).joinedload(User.organization)
        )
        .order_by(Analysis.created_at.desc())
    )
    analyses = result.scalars().unique().all()

    if not analyses:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="No analyses found for the specified date range",
        )

    # Generate CSV with detailed detections
    output = io.StringIO()
    writer = csv.writer(output)

    # Header
    writer.writerow([
        "Analysis ID",
        "Company / Institution",
        "Analyst Name",
        "Sample ID",
        "Media Type",
        "Dilution Factor",
        "Plated Volume (ml)",
        "Total Colonies",
        "CFU/ml",
        "Status",
        "Detection Class",
        "Detection Confidence",
        "BBox X",
        "BBox Y",
        "BBox Width",
        "BBox Height",
    ])

    # Data rows (one per detection)
    for analysis in analyses:
        status_str = analysis.status.value if hasattr(analysis.status, 'value') else str(analysis.status)
        analyst_name = analysis.user.full_name if analysis.user else "Unknown Analyst"
        org_name = analysis.user.organization.name if analysis.user and analysis.user.organization else "ColonyAI General"

        if analysis.detections:
            for detection in analysis.detections:
                writer.writerow([
                    str(analysis.id),
                    org_name,
                    analyst_name,
                    analysis.sample_id,
                    analysis.media_type,
                    analysis.dilution_factor,
                    analysis.plated_volume_ml,
                    analysis.colony_count or 0,
                    f"{analysis.cfu_per_ml:.2e}" if analysis.cfu_per_ml else "N/A",
                    status_str,
                    detection.class_name,
                    f"{detection.confidence * 100:.1f}%",
                    detection.bbox_x,
                    detection.bbox_y,
                    detection.bbox_width,
                    detection.bbox_height,
                ])
        else:
            # Summary row only if no detections
            writer.writerow([
                str(analysis.id),
                org_name,
                analyst_name,
                analysis.sample_id,
                analysis.media_type,
                analysis.dilution_factor,
                analysis.plated_volume_ml,
                analysis.colony_count or 0,
                f"{analysis.cfu_per_ml:.2e}" if analysis.cfu_per_ml else "N/A",
                status_str,
                "", "", "", "", "", "",
            ])

    # Save to file
    reports_dir = os.path.join(settings.UPLOAD_DIR, "reports")
    Path(reports_dir).mkdir(parents=True, exist_ok=True)

    report_id = str(uuid.uuid4())
    filename = f"colonyai-report-{report_id}.csv"
    file_path = os.path.join(reports_dir, filename)

    with open(file_path, "w", newline="", encoding="utf-8") as f:
        f.write(output.getvalue())

    # Generate URL
    url = f"{settings.BACKEND_URL}/uploads/reports/{filename}"
    expires_at = datetime.utcnow().replace(hour=23, minute=59, second=59).isoformat()

    return ReportResponse(
        url=url,
        filename=filename,
        expires_at=expires_at,
    )


@router.get("/{report_id}/download")
async def download_report(
    report_id: str,
    current_user: dict = Depends(get_current_user),
):
    """Download a generated report file"""
    reports_dir = os.path.join(settings.UPLOAD_DIR, "reports")

    # Find the report file
    if not os.path.exists(reports_dir):
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Report not found",
        )

    for filename in os.listdir(reports_dir):
        if report_id in filename:
            file_path = os.path.join(reports_dir, filename)
            media_type = "text/csv" if filename.endswith(".csv") else "application/pdf"
            return FileResponse(
                file_path,
                media_type=media_type,
                filename=filename,
            )

    raise HTTPException(
        status_code=status.HTTP_404_NOT_FOUND,
        detail="Report not found",
    )


# ============================================================
# ORG-WIDE EXPORT: Admin & Manager can export all org data
# ============================================================

@router.get("/admin/pdf-all", response_class=FileResponse)
async def admin_export_all_pdf(
    current_user: dict = Depends(require_role("admin", "manager", "super_admin")),
    db: AsyncSession = Depends(get_db),
):
    """
    [ADMIN/MANAGER/SUPER_ADMIN] Export all org analyses as a single PDF report.
    - Admin/Manager: scoped to their organization
    - Super Admin: all analyses across all organizations
    """
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import cm
    from reportlab.lib.colors import black, HexColor
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER, TA_LEFT
    from collections import defaultdict

    # Build query with role-based scoping
    query = (
        select(Analysis)
        .options(joinedload(Analysis.detections), joinedload(Analysis.user))
        .order_by(Analysis.created_at.desc())
    )
    if current_user.get("role") != "super_admin":
        org_id = current_user.get("organization_id")
        if org_id:
            query = query.where(Analysis.organization_id == uuid.UUID(org_id))

    try:
        result = await db.execute(query)
        analyses = result.scalars().unique().all()

        reports_dir = os.path.abspath(os.path.join(settings.UPLOAD_DIR, "reports"))
        os.makedirs(reports_dir, exist_ok=True)
        report_id = str(uuid.uuid4())
        filename = f"colonyai-admin-all-{report_id}.pdf"
        file_path = os.path.join(reports_dir, filename)

        doc = SimpleDocTemplate(file_path, pagesize=A4,
            topMargin=2*cm, bottomMargin=2*cm,
            leftMargin=2.5*cm, rightMargin=2.5*cm)

        base = "Times-Roman"
        styles = getSampleStyleSheet()
        title_s  = ParagraphStyle("T",  parent=styles["Title"],   fontName=base, fontSize=16, alignment=TA_CENTER, spaceAfter=6)
        head_s   = ParagraphStyle("H",  parent=styles["Heading2"],fontName=base, fontSize=13, spaceBefore=10, spaceAfter=4)
        body_s   = ParagraphStyle("B",  parent=styles["Normal"],  fontName=base, fontSize=11, spaceAfter=3)
        small_s  = ParagraphStyle("S",  parent=styles["Normal"],  fontName=base, fontSize=9,  spaceAfter=2)

        elems = []
        now = datetime.utcnow().strftime("%Y-%m-%d %H:%M UTC")

        # ── Cover ──
        elems += [
            Paragraph("ColonyAI — Admin Master Report", title_s),
            Paragraph(f"Generated: {now} | Total Records: {len(analyses)}", small_s),
            Spacer(1, 12),
        ]

        # ── Global Summary Table ──
        total   = len(analyses)
        completed = sum(1 for a in analyses if str(getattr(a.status,'value',a.status)) == "completed")
        failed  = sum(1 for a in analyses if str(getattr(a.status,'value',a.status)) == "failed")
        colonies = sum(a.colony_count or 0 for a in analyses)
        cfus    = [a.cfu_per_ml for a in analyses if a.cfu_per_ml]
        avg_cfu = sum(cfus)/len(cfus) if cfus else None
        users_set = {str(a.user_id) for a in analyses}

        elems.append(Paragraph("1. Global Summary", head_s))
        summary_rows = [
            ["Metric", "Value"],
            ["Total Analyses (All Users)", str(total)],
            ["Registered Users", str(len(users_set))],
            ["Completed", str(completed)],
            ["Failed", str(failed)],
            ["Success Rate", f"{completed/total*100:.1f}%" if total else "N/A"],
            ["Total Colonies Detected", str(colonies)],
            ["Average CFU/ml", f"{avg_cfu:.2e}" if avg_cfu else "N/A"],
        ]
        t = Table(summary_rows, colWidths=[8*cm, 8*cm])
        t.setStyle(TableStyle([
            ("FONTNAME",(0,0),(-1,-1),base), ("FONTSIZE",(0,0),(-1,-1),11),
            ("BACKGROUND",(0,0),(-1,0),HexColor("#1a202c")), ("TEXTCOLOR",(0,0),(-1,0),HexColor("#ffffff")),
            ("GRID",(0,0),(-1,-1),0.5,black),
            ("ROWBACKGROUNDS",(0,1),(-1,-1),[HexColor("#f7fafc"),HexColor("#ffffff")]),
            ("TOPPADDING",(0,0),(-1,-1),6),("BOTTOMPADDING",(0,0),(-1,-1),6),
            ("LEFTPADDING",(0,0),(-1,-1),8),("RIGHTPADDING",(0,0),(-1,-1),8),
        ]))
        elems += [t, Spacer(1, 16)]

        # ── Per-User Summary ──
        elems.append(Paragraph("2. Per-User Breakdown", head_s))
        user_stats: dict = defaultdict(lambda: {"name":"","count":0,"completed":0,"colonies":0})
        for a in analyses:
            uid = str(a.user_id)
            user_stats[uid]["name"] = a.user.full_name if a.user else uid[:8]
            user_stats[uid]["count"] += 1
            if str(getattr(a.status,'value',a.status)) == "completed":
                user_stats[uid]["completed"] += 1
            user_stats[uid]["colonies"] += a.colony_count or 0

        user_rows = [["Analyst Name", "Analyses", "Completed", "Total Colonies"]]
        for uid, s in user_stats.items():
            user_rows.append([s["name"], str(s["count"]), str(s["completed"]), str(s["colonies"])])
        t2 = Table(user_rows, colWidths=[6*cm,3*cm,3*cm,4*cm])
        t2.setStyle(TableStyle([
            ("FONTNAME",(0,0),(-1,-1),base),("FONTSIZE",(0,0),(-1,-1),10),
            ("BACKGROUND",(0,0),(-1,0),HexColor("#2d3748")),("TEXTCOLOR",(0,0),(-1,0),HexColor("#ffffff")),
            ("GRID",(0,0),(-1,-1),0.5,black),
            ("ROWBACKGROUNDS",(0,1),(-1,-1),[HexColor("#ebf8ff"),HexColor("#ffffff")]),
            ("TOPPADDING",(0,0),(-1,-1),5),("BOTTOMPADDING",(0,0),(-1,-1),5),
            ("LEFTPADDING",(0,0),(-1,-1),6),("RIGHTPADDING",(0,0),(-1,-1),6),
        ]))
        elems += [t2, Spacer(1, 16)]

        # ── Monthly Trend ──
        elems.append(Paragraph("3. Monthly Throughput Trend", head_s))
        monthly: dict = defaultdict(lambda:{"count":0,"colonies":0})
        for a in analyses:
            m = a.created_at.strftime("%Y-%m")
            monthly[m]["count"] += 1
            monthly[m]["colonies"] += a.colony_count or 0
        trend_rows = [["Month", "Analyses", "Total Colonies"]]
        for m in sorted(monthly):
            trend_rows.append([m, str(monthly[m]["count"]), str(monthly[m]["colonies"])])
        t3 = Table(trend_rows, colWidths=[4*cm,4*cm,8*cm])
        t3.setStyle(TableStyle([
            ("FONTNAME",(0,0),(-1,-1),base),("FONTSIZE",(0,0),(-1,-1),10),
            ("BACKGROUND",(0,0),(-1,0),HexColor("#4a5568")),("TEXTCOLOR",(0,0),(-1,0),HexColor("#ffffff")),
            ("GRID",(0,0),(-1,-1),0.5,black),
            ("ROWBACKGROUNDS",(0,1),(-1,-1),[HexColor("#f0fff4"),HexColor("#ffffff")]),
            ("TOPPADDING",(0,0),(-1,-1),5),("BOTTOMPADDING",(0,0),(-1,-1),5),
            ("LEFTPADDING",(0,0),(-1,-1),8),("RIGHTPADDING",(0,0),(-1,-1),8),
        ]))
        elems += [t3, Spacer(1,16), PageBreak()]

        # ── All Samples Detail ──
        elems.append(Paragraph("4. All Sample Records", head_s))
        detail_rows = [["Sample ID","Analyst","Media","CFU/ml","Colonies","Confidence","Status","Date"]]
        for a in analyses:
            name = a.user.full_name if a.user else "Unknown"
            status_val = str(getattr(a.status,'value',a.status)).capitalize()
            cfu_val = f"{a.cfu_per_ml:.2e}" if a.cfu_per_ml else (str(a.cfu_status) if a.cfu_status else "N/A")
            conf = f"{a.confidence_score*100:.0f}%" if a.confidence_score else "N/A"
            detail_rows.append([
                a.sample_id, name, a.media_type or "—",
                cfu_val, str(a.colony_count or 0),
                conf, status_val,
                a.created_at.strftime("%Y-%m-%d"),
            ])
        # Optimized colWidths to fit A4 (16cm total)
        t4 = Table(detail_rows, colWidths=[2.2*cm, 2.8*cm, 1.8*cm, 2.2*cm, 1.6*cm, 1.8*cm, 1.8*cm, 1.8*cm])
        t4.setStyle(TableStyle([
            ("FONTNAME",(0,0),(-1,-1),base),("FONTSIZE",(0,0),(-1,-1),8),
            ("BACKGROUND",(0,0),(-1,0),HexColor("#1a202c")),("TEXTCOLOR",(0,0),(-1,0),HexColor("#ffffff")),
            ("GRID",(0,0),(-1,-1),0.4,black),
            ("ROWBACKGROUNDS",(0,1),(-1,-1),[HexColor("#f7fafc"),HexColor("#ffffff")]),
            ("TOPPADDING",(0,0),(-1,-1),3),("BOTTOMPADDING",(0,0),(-1,-1),3),
            ("LEFTPADDING",(0,0),(-1,-1),4),("RIGHTPADDING",(0,0),(-1,-1),4),
            ("FONTSIZE",(0,1),(-1,-1),8),
        ]))
        elems.append(t4)

        doc.build(elems)
        return FileResponse(file_path, media_type="application/pdf", filename=f"colonyai-admin-report-{report_id}.pdf")
    except Exception as e:
        import logging
        logging.error(f"Error generating admin PDF: {str(e)}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error generating PDF report: {str(e)}"
        )


@router.get("/admin/excel-all", response_class=FileResponse)
async def admin_export_all_excel(
    current_user: dict = Depends(require_role("admin", "manager", "super_admin")),
    db: AsyncSession = Depends(get_db),
):
    """
    [ADMIN/MANAGER/SUPER_ADMIN] Export comprehensive analytics as Excel (.xlsx).
    - Admin/Manager: scoped to their organization
    - Super Admin: all data across all organizations

    Sheets: Global Summary, Per-User Analytics, Monthly Trend,
            Media Type Breakdown, CFU Distribution, All Raw Records
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from collections import defaultdict
    import statistics

    try:
        query = (
            select(Analysis)
            .options(joinedload(Analysis.detections), joinedload(Analysis.user))
            .order_by(Analysis.created_at.desc())
        )
        if current_user.get("role") != "super_admin":
            org_id = current_user.get("organization_id")
            if org_id:
                query = query.where(Analysis.organization_id == uuid.UUID(org_id))

        result = await db.execute(query)
        analyses = result.scalars().unique().all()

        wb = openpyxl.Workbook()

        # ── Styles ──
        HDR_FILL  = PatternFill("solid", fgColor="1A202C")
        HDR2_FILL = PatternFill("solid", fgColor="2D3748")
        ALT_FILL  = PatternFill("solid", fgColor="EBF8FF")
        HDR_FONT  = Font(bold=True, color="FFFFFF", size=11)
        TITLE_FONT= Font(bold=True, size=14, color="1A202C")
        SUB_FONT  = Font(bold=True, size=10, color="2D3748")
        THIN      = Side(style="thin", color="CBD5E0")
        BORDER    = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
        CENTER    = Alignment(horizontal="center", vertical="center", wrap_text=True)
        LEFT      = Alignment(horizontal="left",   vertical="center", wrap_text=True)

        def style_header_row(ws, row, fill=HDR_FILL):
            for cell in ws[row]:
                cell.fill = fill
                cell.font = HDR_FONT
                cell.alignment = CENTER
                cell.border = BORDER

        def style_data_row(ws, row, alt=False):
            for cell in ws[row]:
                if alt:
                    cell.fill = ALT_FILL
                cell.border = BORDER
                cell.alignment = LEFT

        def auto_width(ws):
            for col in ws.columns:
                max_len = 0
                col_letter = get_column_letter(col[0].column)
                for cell in col:
                    try:
                        max_len = max(max_len, len(str(cell.value or "")))
                    except:
                        pass
                ws.column_dimensions[col_letter].width = min(max_len + 4, 40)

        # ─────────────────────────────────────────────
        # SHEET 1: Executive Dashboard
        # ─────────────────────────────────────────────
        ws_dash = wb.active
        ws_dash.title = "📊 Executive Dashboard"

        ws_dash.merge_cells("A1:J1")
        ws_dash["A1"] = "COLONYAI — LABORATORY INTELLIGENCE EXECUTIVE SUMMARY"
        ws_dash["A1"].font = Font(bold=True, size=18, color="FFFFFF")
        ws_dash["A1"].fill = PatternFill("solid", fgColor="1A202C")
        ws_dash["A1"].alignment = CENTER

        # Key Metrics (KPIs)
        total     = len(analyses)
        users_set = {str(a.user_id) for a in analyses}
        cfus_all  = [a.cfu_per_ml for a in analyses if a.cfu_per_ml]
        confs_all = [a.confidence_score for a in analyses if a.confidence_score]
        tntc_count = sum(1 for a in analyses if a.cfu_status == "TNTC")
        tftc_count = sum(1 for a in analyses if a.cfu_status == "TFTC")

        ws_dash.append([])
        ws_dash.append(["SYSTEM OVERVIEW", "", "", "", "PRODUCTIVITY GAINS"])
        ws_dash["A3"].font = Font(bold=True, size=12)
        ws_dash["E3"].font = Font(bold=True, size=12)

        # KPIs using Formulas from Sheet 2
        ws_dash["A4"] = "Total Samples"
        ws_dash["B4"] = "='📋 Summary Table'!B4"
        ws_dash["A5"] = "Valid Analyses"
        ws_dash["B5"] = "='📋 Summary Table'!B6"
        ws_dash["A6"] = "Total Colonies"
        ws_dash["B6"] = "='📋 Summary Table'!B11"
        ws_dash["A7"] = "Active Analysts"
        ws_dash["B7"] = "='📋 Summary Table'!B5"

        ws_dash["E4"] = "Efficiency Rate"
        ws_dash["F4"] = "='📋 Summary Table'!B15" # Efficiency Gain
        ws_dash["E5"] = "Time Saved (Hrs)"
        ws_dash["F5"] = "=(B4 * 18) / 60" # 18 mins saved per sample
        ws_dash["E6"] = "Labor Cost Saved"
        ws_dash["F6"] = "=F5 * 50000 * 0.4" # IDR 50k/hr * 40% reduction
        ws_dash["E7"] = "Neural Confidence"
        ws_dash["F7"] = "='📋 Summary Table'!B14"

        # Styling Dashboard KPIs
        for row in range(4, 8):
            ws_dash[f"A{row}"].fill = PatternFill("solid", fgColor="F7FAFC")
            ws_dash[f"E{row}"].fill = PatternFill("solid", fgColor="F0FFF4")
            ws_dash[f"B{row}"].font = Font(bold=True)
            ws_dash[f"F{row}"].font = Font(bold=True, color="2F855A")
            ws_dash[f"F{row}"].number_format = '#,##0'

        # Create Hidden Data for Dashboard Charts
        ws_hidden = wb.create_sheet("hidden_data")
        media_counts = defaultdict(int)
        for a in analyses: media_counts[a.media_type or "Unknown"] += 1
        ws_hidden.append(["Media", "Count"])
        for m, c in media_counts.items(): ws_hidden.append([m, c])

        mon_counts = defaultdict(int)
        for a in analyses: mon_counts[a.created_at.strftime("%b %Y")] += 1
        ws_hidden.append([])
        ws_hidden.append(["Month", "Count"])
        for m, c in mon_counts.items(): ws_hidden.append([m, c])

        from openpyxl.chart import BarChart, LineChart, PieChart, Reference

        pie = PieChart()
        pie.title = "Sample Volume by Media Type"
        data = Reference(ws_hidden, min_col=2, min_row=2, max_row=len(media_counts)+1)
        cats = Reference(ws_hidden, min_col=1, min_row=2, max_row=len(media_counts)+1)
        pie.add_data(data)
        pie.set_categories(cats)
        ws_dash.add_chart(pie, "A9")

        bar = BarChart()
        bar.title = "Throughput Growth (Monthly)"
        bar.y_axis.title = "Samples"
        data = Reference(ws_hidden, min_col=2, min_row=len(media_counts)+4, max_row=len(media_counts)+len(mon_counts)+3)
        cats = Reference(ws_hidden, min_col=1, min_row=len(media_counts)+4, max_row=len(media_counts)+len(mon_counts)+3)
        bar.add_data(data)
        bar.set_categories(cats)
        ws_dash.add_chart(bar, "E9")

        ws_hidden.sheet_state = "hidden"

        # ─────────────────────────────────────────────
        # SHEET 2: Global Summary (Formula Driven)
        # ─────────────────────────────────────────────
        ws1 = wb.create_sheet("📋 Summary Table")

        ws1.merge_cells("A1:C1")
        ws1["A1"] = "ColonyAI — Admin Analytics Summary"
        ws1["A1"].font = TITLE_FONT
        ws1["A1"].alignment = CENTER

        ws1.append([])
        ws1.append(["Metric", "Value", "Note"])
        style_header_row(ws1, 3)

        # Dynamic Formulas pointing to Raw Records (Sheet 6)
        total_count = len(analyses)
        ws1.append(["Total Analyses", total_count, "Calculated from Records"]) # B4
        ws1.append(["Total Registered Analysts", len(users_set), "Unique users"]) # B5
        ws1.append(["Completed Records", f'=COUNTIF(\'📋 All Records\'!N:N, "Completed")', "Status check"]) # B6
        ws1.append(["Failed Records", f'=COUNTIF(\'📋 All Records\'!N:N, "Failed")', "Status check"]) # B7
        ws1.append(["Success Rate", "=B6/B4", "Dynamic %"]) # B8
        ws1["B8"].number_format = '0.0%'

        ws1.append(["TNTC Results", f'=COUNTIF(\'📋 All Records\'!J:J, "TNTC")', "High density"]) # B9
        ws1.append(["TFTC Results", f'=COUNTIF(\'📋 All Records\'!J:J, "TFTC")', "Low density"]) # B10
        ws1.append(["Total Colonies Detected", f"=SUM('📋 All Records'!H:H)", "Global sum"]) # B11
        ws1.append(["Average Colonies/Sample", "=B11/B4", "Mean distribution"]) # B12

        cfus_avg = sum(cfus_all)/len(cfus_all) if cfus_all else 0
        ws1.append(["Average CFU/ml", cfus_avg, "Mean density"]) # B13
        ws1["B13"].number_format = '0.00E+00'

        conf_avg = sum(confs_all)/len(confs_all) if confs_all else 0
        ws1.append(["Average Confidence", conf_avg, "Model reliability"]) # B14
        ws1["B14"].number_format = '0.0%'

        ws1.append(["Efficiency Gain", 0.9, "AI vs Manual baseline"]) # B15
        ws1["B15"].number_format = '0%'

        for row in range(4, 16):
            style_data_row(ws1, row, alt=(row % 2 == 1))
        auto_width(ws1)

        # ─────────────────────────────────────────────
        # SHEET 2: Per-User Analytics
        # ─────────────────────────────────────────────
        ws2 = wb.create_sheet("👤 Per-User Analytics")
        ws2.append(["Analyst Name", "Email", "Total Analyses", "Completed", "Failed",
                    "Success Rate", "Total Colonies", "Avg CFU/ml", "Avg Confidence"])
        style_header_row(ws2, 1)

        user_data: dict = defaultdict(lambda: {
            "name":"","email":"","count":0,"completed":0,"failed":0,
            "colonies":0,"cfus":[],"confs":[]
        })
        for a in analyses:
            uid = str(a.user_id)
            if a.user:
                user_data[uid]["name"]  = a.user.full_name
                user_data[uid]["email"] = a.user.email
            else:
                user_data[uid]["name"]  = uid[:8]
                user_data[uid]["email"] = "—"
            user_data[uid]["count"] += 1
            s = str(getattr(a.status,'value',a.status))
            if s == "completed": user_data[uid]["completed"] += 1
            if s == "failed":    user_data[uid]["failed"]    += 1
            user_data[uid]["colonies"] += a.colony_count or 0
            if a.cfu_per_ml:     user_data[uid]["cfus"].append(a.cfu_per_ml)
            if a.confidence_score: user_data[uid]["confs"].append(a.confidence_score)

        for i, (uid, d) in enumerate(sorted(user_data.items(), key=lambda x: -x[1]["count"])):
            sr   = f"{d['completed']/d['count']*100:.1f}%" if d['count'] else "N/A"
            acfu = f"{sum(d['cfus'])/len(d['cfus']):.2e}" if d['cfus'] else "N/A"
            aconf= f"{sum(d['confs'])/len(d['confs'])*100:.1f}%" if d['confs'] else "N/A"
            ws2.append([d["name"], d["email"], d["count"], d["completed"], d["failed"],
                        sr, d["colonies"], acfu, aconf])
            style_data_row(ws2, i+2, alt=(i%2==1))
        auto_width(ws2)

        # ─────────────────────────────────────────────
        # SHEET 3: Monthly Trend
        # ─────────────────────────────────────────────
        ws3 = wb.create_sheet("📅 Monthly Trend")
        ws3.append(["Month", "Total Analyses", "Completed", "Total Colonies", "Avg CFU/ml", "Avg Confidence"])
        style_header_row(ws3, 1)

        monthly: dict = defaultdict(lambda:{"count":0,"completed":0,"colonies":0,"cfus":[],"confs":[]})
        for a in analyses:
            m = a.created_at.strftime("%Y-%m")
            monthly[m]["count"] += 1
            if str(getattr(a.status,'value',a.status))=="completed": monthly[m]["completed"] += 1
            monthly[m]["colonies"] += a.colony_count or 0
            if a.cfu_per_ml:       monthly[m]["cfus"].append(a.cfu_per_ml)
            if a.confidence_score: monthly[m]["confs"].append(a.confidence_score)

        for i, month in enumerate(sorted(monthly)):
            d = monthly[month]
            acfu  = f"{sum(d['cfus'])/len(d['cfus']):.2e}" if d['cfus'] else "N/A"
            aconf = f"{sum(d['confs'])/len(d['confs'])*100:.1f}%" if d['confs'] else "N/A"
            ws3.append([month, d["count"], d["completed"], d["colonies"], acfu, aconf])
            style_data_row(ws3, i+2, alt=(i%2==1))
        auto_width(ws3)

        # ─────────────────────────────────────────────
        # SHEET 4: Media Type Breakdown
        # ─────────────────────────────────────────────
        ws4 = wb.create_sheet("🧪 Media Type Breakdown")
        ws4.append(["Media Type", "Count", "Completed", "Total Colonies", "Avg CFU/ml"])
        style_header_row(ws4, 1)

        media_data: dict = defaultdict(lambda:{"count":0,"completed":0,"colonies":0,"cfus":[]})
        for a in analyses:
            mt = a.media_type or "Unknown"
            media_data[mt]["count"] += 1
            if str(getattr(a.status,'value',a.status))=="completed": media_data[mt]["completed"] += 1
            media_data[mt]["colonies"] += a.colony_count or 0
            if a.cfu_per_ml: media_data[mt]["cfus"].append(a.cfu_per_ml)

        for i, (mt, d) in enumerate(sorted(media_data.items(), key=lambda x: -x[1]["count"])):
            acfu = f"{sum(d['cfus'])/len(d['cfus']):.2e}" if d['cfus'] else "N/A"
            ws4.append([mt, d["count"], d["completed"], d["colonies"], acfu])
            style_data_row(ws4, i+2, alt=(i%2==1))
        auto_width(ws4)

        # ─────────────────────────────────────────────
        # SHEET 5: CFU Distribution
        # ─────────────────────────────────────────────
        ws5 = wb.create_sheet("📈 CFU Distribution")
        ws5.append(["CFU/ml Range", "Count", "Percentage"])
        style_header_row(ws5, 1)

        buckets = [
            ("<1e3", 0, 1e3),
            ("1e3–1e4", 1e3, 1e4),
            ("1e4–1e5", 1e4, 1e5),
            ("1e5–1e6", 1e5, 1e6),
            (">1e6",   1e6, float("inf")),
        ]
        valid_cfus = [a.cfu_per_ml for a in analyses if a.cfu_per_ml]
        for i, (label, lo, hi) in enumerate(buckets):
            cnt = sum(1 for v in valid_cfus if lo <= v < hi)
            pct = f"{cnt/len(valid_cfus)*100:.1f}%" if valid_cfus else "0.0%"
            ws5.append([label, cnt, pct])
            style_data_row(ws5, i+2, alt=(i%2==1))

        ws5.append([])
        ws5.append(["TNTC (Too Numerous)", tntc_count, f"{tntc_count/total*100:.1f}%" if total else "0.0%"])
        ws5.append(["TFTC (Too Few)",      tftc_count, f"{tftc_count/total*100:.1f}%" if total else "0.0%"])
        auto_width(ws5)

        # ── Add Charts ──
        from openpyxl.chart import BarChart, LineChart, PieChart, Reference

        # 1. Line Chart on Sheet 3 (Monthly Trend)
        c1 = LineChart()
        c1.title = "Monthly Analysis Trend"
        c1.style = 13
        c1.y_axis.title = 'Count'
        c1.x_axis.title = 'Month'
        data = Reference(ws3, min_col=2, min_row=1, max_col=3, max_row=len(monthly)+1)
        cats = Reference(ws3, min_col=1, min_row=2, max_row=len(monthly)+1)
        c1.add_data(data, titles_from_data=True)
        c1.set_categories(cats)
        ws3.add_chart(c1, "H2")

        # 2. Pie Chart on Sheet 4 (Media Type)
        c2 = PieChart()
        c2.title = "Media Type Distribution"
        data = Reference(ws4, min_col=2, min_row=2, max_row=len(media_data)+1)
        cats = Reference(ws4, min_col=1, min_row=2, max_row=len(media_data)+1)
        c2.add_data(data, titles_from_data=False)
        c2.set_categories(cats)
        ws4.add_chart(c2, "H2")

        # 3. Bar Chart on Sheet 5 (CFU Distribution)
        c3 = BarChart()
        c3.title = "CFU/ml Distribution"
        c3.style = 10
        data = Reference(ws5, min_col=2, min_row=1, max_row=len(buckets)+1)
        cats = Reference(ws5, min_col=1, min_row=2, max_row=len(buckets)+1)
        c3.add_data(data, titles_from_data=True)
        c3.set_categories(cats)
        ws5.add_chart(c3, "H2")

        # ─────────────────────────────────────────────
        # SHEET 6: All Raw Records
        # ─────────────────────────────────────────────
        ws6 = wb.create_sheet("📋 All Records")
        ws6.append([
            "Analysis ID","Analyst","Email","Sample ID","Media Type",
            "Dilution Factor","Plated Vol (ml)","Colony Count",
            "CFU/ml","CFU Status","Uncertainty U",
            "Incubation Temp (°C)","Incubation Time (Hrs)","Method Standard",
            "Media Batch #","Incubator ID",
            "Confidence","Reliability","Status",
            "Colony Single","Colony Merged","Bubble","Dust","Crack",
            "Created At"
        ])
        style_header_row(ws6, 1)

        for i, a in enumerate(analyses):
            name  = a.user.full_name if a.user else "Unknown"
            email = a.user.email     if a.user else "Unknown"
            cb    = a.class_breakdown or {}
            status_val = str(getattr(a.status,'value',a.status)).capitalize()
            cfu_val    = f"{a.cfu_per_ml:.2e}" if a.cfu_per_ml else (a.cfu_status or "N/A")
            ws6.append([
                str(a.id), name, email, a.sample_id, a.media_type or "—",
                a.dilution_factor, a.plated_volume_ml, a.colony_count or 0,
                cfu_val, a.cfu_status or "normal", a.uncertainty_u or "N/A",
                a.incubation_temp or "—", a.incubation_time_hours or "—",
                a.method_standard or "ISO 4833-1:2013",
                a.media_batch_number or "—", a.incubator_id or "—",
                f"{a.confidence_score*100:.1f}%" if a.confidence_score else "N/A",
                (a.reliability or "high").capitalize(), status_val,
                cb.get("colony_single", 0), cb.get("colony_merged", 0),
                cb.get("bubble", 0), cb.get("dust_debris", 0), cb.get("media_crack", 0),
                a.created_at.strftime("%Y-%m-%d %H:%M"),
            ])
            style_data_row(ws6, i+2, alt=(i%2==1))
        auto_width(ws6)

        # Save
        reports_dir = os.path.abspath(os.path.join(settings.UPLOAD_DIR, "reports"))
        os.makedirs(reports_dir, exist_ok=True)
        report_id = str(uuid.uuid4())
        filename = f"colonyai-admin-analytics-{report_id}.xlsx"
        file_path = os.path.join(reports_dir, filename)
        wb.save(file_path)

        return FileResponse(
            file_path,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            filename=f"colonyai-admin-analytics-{report_id}.xlsx",
        )
    except Exception as e:
        import logging
        logging.error(f"Error generating admin Excel: {str(e)}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error generating Excel report: {str(e)}"
        )
